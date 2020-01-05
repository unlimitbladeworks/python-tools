# -*- coding: utf-8 -*-
"""
@Author  : Sy
@File    : glory_excel.py
@Time    : 2020-01-05 16:57
@desc    : 处理最终的 excel
"""

from openpyxl.drawing.image import Image
from openpyxl import *
import requests
from PIL import Image as PILImage

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko)\
    Chrome/69.0.3497.100 Safari/537.36'
}


class Excel(object):
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)  # 加载，可读
        sheets = self.wb.sheetnames  # 获取所有sheet页
        self.sheet = sheets[0]  # 默认第一sheet页
        self.ws = self.wb[self.sheet]  # 切换到sheet页

    # 获取表格的总行数和总列数
    def get_rows_cols_num(self):
        rows = self.ws.max_row  # 最大行数
        columns = self.ws.max_column  # 最小行数
        return rows, columns

    # 获取某个单元格的值
    def get_cell_values(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value  # 设置值
        return cell_value

    # 获取某列的所有值
    def get_col_values(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_values(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, column, cell_value):
        try:
            if isinstance(cell_value, Image):
                self.ws.add_image(cell_value, f'{row}{column}')
            else:
                self.ws.cell(row=row, column=column).value = cell_value
        except Exception as e:
            self.ws.cell(row=row, column=column).value = "None"

    # 设置单元格宽度和高度
    def set_cell_height(self, row, col, row_height, column_weight):
        self.ws.row_dimensions[row].height = row_height
        self.ws.column_dimensions[col].width = column_weight

    # 保存文件
    def save_file(self, file_name=None):
        if file_name:
            self.wb.save(file_name)
        else:
            self.wb.save(self.file)


def http_client(url):
    """ 请求 """
    r = requests.get(url=url, headers=headers)
    if r.status_code == 200:
        return r.content
    else:
        return None


def write_img(filename, content):
    """ 二进制图片写入文件中 """
    with open(filename, 'wb') as f:
        f.write(content)


def main():
    try:
        excel = Excel('glory.xlsx')
        avatar_list = excel.get_col_values(3)  # 获取 C 列，头像url地址
        for index, value in enumerate(avatar_list):
            if 'http' not in value:
                continue
            img_path = f'./img/{index}.png'
            img = http_client(value)
            write_img(img_path, img)  # 写入当前路径下 img 目录下图片
            img = PILImage.open(img_path)
            image = img.resize((75, 75), PILImage.ANTIALIAS)  # 重构图片大小，设置为 75 像素
            image.save(img_path)
            picture = Image(img_path)
            excel.set_cell_height(index + 1, 'D', 59, 11.15)  # D 列头像列，设置行宽，59，11.15
            excel.set_cell_value('D', index + 1, picture)

        excel.save_file('glory_new.xlsx')  # 保存文件
    except Exception as e:
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
